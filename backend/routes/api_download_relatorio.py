import io
import calendar
import re
from datetime import datetime, timedelta
from flask import Blueprint, send_file, request, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from models.database import Relatorios

Download_Relatorio = Blueprint("Download_Relatorio", __name__)

def limpar_e_converter_float(valor):
    """Função de segurança para limpar qualquer string suja do banco e converter para float"""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        # Remove símbolos, 'AOA', espaços e padroniza o ponto decimal
        valor_limpo = re.sub(r'[^\d,.-]', '', str(valor)).strip()
        if ',' in valor_limpo and '.' in valor_limpo:
            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
        elif ',' in valor_limpo:
            valor_limpo = valor_limpo.replace(',', '.')
        return float(valor_limpo)
    except Exception:
        return 0.0

def limpar_e_converter_int(valor):
    """Função de segurança para converter a quantidade para inteiro"""
    if valor is None:
        return 0
    try:
        return int(float(str(valor).replace(',', '.')))
    except Exception:
        return 0

@Download_Relatorio.route("/exportar_excel/<int:id_empresario>", methods=["GET"])
def exportar_excel(id_empresario):
    try:
        # 1. Filtros de Dados baseados no Frontend
        periodo = request.args.get("periodo", "todos")
        mes_especifico = request.args.get("mes", type=int)
        ano_especifico = request.args.get("ano", type=int)

        hoje = datetime.utcnow()
        query = Relatorios.query.filter(Relatorios.id_empresario == id_empresario)

        if mes_especifico and ano_especifico:
            inicio_periodo = datetime(ano_especifico, mes_especifico, 1, 0, 0, 0)
            ultimo_dia = calendar.monthrange(ano_especifico, mes_especifico)[1]
            fim_periodo = datetime(ano_especifico, mes_especifico, ultimo_dia, 23, 59, 59)
            query = query.filter(Relatorios.data_venda_produto >= inicio_periodo, Relatorios.data_venda_produto <= fim_periodo)
        elif periodo == "hoje":
            query = query.filter(Relatorios.data_venda_produto >= hoje.replace(hour=0, minute=0, second=0, microsecond=0))
        elif periodo == "semana":
            inicio_semana = hoje - timedelta(days=hoje.weekday())
            query = query.filter(Relatorios.data_venda_produto >= inicio_semana.replace(hour=0, minute=0, second=0, microsecond=0))
        elif periodo == "mes":
            query = query.filter(Relatorios.data_venda_produto >= hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0))
        elif periodo == "ano":
            query = query.filter(Relatorios.data_venda_produto >= hoje.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0))

        # Ordenação Ascendente para manter a consistência cronológica nos gráficos
        relatorio_das_vendas = query.order_by(Relatorios.data_venda_produto.asc()).all()

        if not relatorio_das_vendas:
            return jsonify({"resposta": "Nenhum dado encontrado para exportar"}), 404

        # 2. Inicialização do Livro e Folha de Cálculo
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        
        # CORREÇÃO DA LINHA DE GRADE: Abordagem direta e universal do openpyxl
        ws.sheet_view.showGridLines = True  

        # Identidade Visual Premium (Paleta Teal Âncora)
        TEAL_DARK = "005B60"
        TEAL_LIGHT = "E0F2F1"
        GRAY_LIGHT = "F4F9F9"
        GRAY_BORDER = "E0E0E0"
        
        font_title = Font(name="Segoe UI", size=15, bold=True, color=TEAL_DARK)
        font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="555555")
        font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="777777")
        font_kpi_val = Font(name="Segoe UI", size=13, bold=True, color=TEAL_DARK)
        font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10, color="333333")
        font_total = Font(name="Segoe UI", size=10, bold=True, color="000000")

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        border_soft = Border(
            left=Side(style="thin", color=GRAY_BORDER), right=Side(style="thin", color=GRAY_BORDER),
            top=Side(style="thin", color=GRAY_BORDER), bottom=Side(style="thin", color=GRAY_BORDER)
        )

        # Cabeçalho Principal do Painel
        ws["B2"] = "ANCORA CONTROL - DASHBOARD DE VENDAS"
        ws["B2"].font = font_title
        ws["B3"] = f"Análise estática resumida | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["B3"].font = font_subtitle

        total_de_dados = len(relatorio_das_vendas)

        # ==========================================================
        # 3. CARDS DE KPI NO TOPO (MÉTRICAS DINÂMICAS)
        # ==========================================================
        # Card 1: Faturamento Bruto Total
        ws.merge_cells("B5:C5")
        ws.merge_cells("B6:C6")
        ws["B5"] = "FATURAMENTO BRUTO TOTAL"
        ws["B6"] = f"=SUM(E10:E{total_de_dados + 9})"
        
        # Card 2: Lucro Líquido Acumulado (Somatório dos Lucros)
        ws.merge_cells("D5:E5")
        ws.merge_cells("D6:E6")
        ws["D5"] = "LUCRO LÍQUIDO ACUMULADO"
        ws["D6"] = f"=SUM(F10:F{total_de_dados + 9})"

        # Aplicar estilos nos blocos de KPI
        for lbl_coord, val_coord in [("B5", "B6"), ("D5", "D6")]:
            ws[lbl_coord].font = font_kpi_label
            ws[lbl_coord].alignment = align_center
            ws[lbl_coord].fill = PatternFill(start_color=GRAY_LIGHT, fill_type="solid")
            ws[val_coord].font = font_kpi_val
            ws[val_coord].alignment = align_center
            ws[val_coord].fill = PatternFill(start_color=TEAL_LIGHT, fill_type="solid")
            ws[val_coord].number_format = '"AOA " #,##0.00'

        # ==========================================================
        # 4. CONSTRUÇÃO DA TABELA DE DADOS PRINCIPAL
        # ==========================================================
        start_row = 9
        headers = ["DATA", "PRODUTO", "QUANTIDADE", "TOTAL (AOA)", "LUCRO (AOA)"]
        
        # Gerar os cabeçalhos (Colunas B a F)
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=TEAL_DARK, fill_type="solid")
            cell.font = font_header
            cell.alignment = align_center

        # Processar e injetar dados (Garantindo limpeza absoluta de strings para evitar quebra)
        for idx, lista in enumerate(relatorio_das_vendas):
            current_row = start_row + 1 + idx
            data_formatada = lista.data_venda_produto.strftime("%Y-%m-%d %H:%M") if hasattr(lista.data_venda_produto, 'strftime') else str(lista.data_venda_produto)
            
            row_data = [
                data_formatada, 
                str(lista.nome_produto_venda), 
                limpar_e_converter_int(lista.quanidade_produto_venda), 
                limpar_e_converter_float(lista.total_pago_produto_venda), 
                limpar_e_converter_float(lista.lucro_produto_venda)
            ]
            
            for col_offset, val in enumerate(row_data, start=2):
                cell = ws.cell(row=current_row, column=col_offset, value=val)
                cell.font = font_body
                cell.border = border_soft
                
                # Efeito Zebra sutil nas linhas alternadas
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color=GRAY_LIGHT, fill_type="solid")
                
                # Alinhamentos e Formatações de Célula
                if col_offset in [2, 4]:  # Data e Quantidade
                    cell.alignment = align_center
                elif col_offset in [5, 6]:  # Valores Monetários
                    cell.number_format = '"AOA " #,##0.00'  
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

        # ==========================================================
        # 5. LINHA INFERIOR DE TOTAL GERAL (SOMA REAL DA PLANILHA)
        # ==========================================================
        total_row = start_row + total_de_dados + 1
        
        ws.cell(row=total_row, column=2, value="TOTAL GERAL").font = font_total
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
        
        # Injeção das Fórmulas de Somatório Nativo
        cell_qtd = ws.cell(row=total_row, column=4, value=f"=SUM(D10:D{total_row-1})")
        cell_total = ws.cell(row=total_row, column=5, value=f"=SUM(E10:E{total_row-1})")
        cell_lucro = ws.cell(row=total_row, column=6, value=f"=SUM(F10:F{total_row-1})")

        # Formatação das Moedas dos Totais Gerais
        cell_total.number_format = '"AOA " #,##0.00'
        cell_lucro.number_format = '"AOA " #,##0.00'

        # Estilização Contábil Profissional da Linha de Encerramento
        for c_idx in range(2, 7):
            cell = ws.cell(row=total_row, column=c_idx)
            cell.font = font_total
            cell.fill = PatternFill(start_color=TEAL_LIGHT, fill_type="solid")
            cell.border = Border(top=Side(style="thin", color="000000"), bottom=Side(style="double", color="000000"))
            if c_idx >= 4:
                cell.alignment = align_right if c_idx > 4 else align_center

        # ==========================================================
        # 6. GRÁFICOS CONFIGURADOS COM PROPORÇÃO COMPACTA (COLUNA H)
        # ==========================================================
        CHART_HEIGHT = 7
        CHART_WIDTH = 11

        # Definição das referências de colunas
        cats_ref = Reference(ws, min_col=2, min_row=start_row+1, max_row=total_row-1)
        quant_ref = Reference(ws, min_col=4, min_row=start_row, max_row=total_row-1)
        money_ref = Reference(ws, min_col=5, min_row=start_row, max_col=6, max_row=total_row-1)

        # Gráfico 1: Barras -> Alocado na célula H5
        chart_bar = BarChart()
        chart_bar.type = "col"
        chart_bar.title = "Faturamento vs Lucro Líquido"
        chart_bar.add_data(money_ref, titles_from_data=True)
        chart_bar.set_categories(cats_ref)
        chart_bar.height = CHART_HEIGHT
        chart_bar.width = CHART_WIDTH
        ws.add_chart(chart_bar, "H5")

        # Gráfico 2: Linha -> Alocado na célula H16
        chart_line = LineChart()
        chart_line.title = "Tendência de Crescimento"
        chart_line.add_data(money_ref, titles_from_data=True)
        chart_line.set_categories(cats_ref)
        chart_line.height = CHART_HEIGHT
        chart_line.width = CHART_WIDTH
        ws.add_chart(chart_line, "H16")

        # Gráfico 3: Pizza -> Alocado na célula H27
        chart_pie = PieChart()
        chart_pie.title = "Volume de Produtos Vendidos"
        chart_pie.add_data(quant_ref, titles_from_data=True)
        chart_pie.set_categories(Reference(ws, min_col=3, min_row=start_row+1, max_row=total_row-1))
        chart_pie.height = CHART_HEIGHT
        chart_pie.width = CHART_WIDTH
        ws.add_chart(chart_pie, "H27")

        # ==========================================================
        # 7. REDIMENSIONAMENTO AUTOMÁTICO DAS COLUNAS DA TABELA
        # ==========================================================
        for col in ws.columns:
            if col[0].column >= 2 and col[0].column <= 6:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # 8. Salvar binário em memória e disparar stream HTTP para download instantâneo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        data_hoje = datetime.now().strftime("%Y-%m-%d")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Dashboard_Ancora_{data_hoje}.xlsx"
        )

    except Exception as erro:
        print(f"Erro fatal na geração do dashboard do Excel: {erro}")
        return jsonify({"resposta": f"Erro interno do servidor ao gerar arquivo: {str(erro)}"}), 500